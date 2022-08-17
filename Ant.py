import os
from telnetlib import Telnet
from openpyxl import Workbook
from netmiko import ConnectHandler


class AutoNetT:
    def __init__(self, Workbook, is_Admin, user):
        self.Telnet = None
        self.host = input()
        self.user = user
        self.is_Admin = is_Admin
        self.workbook = Workbook()
        self.path = os.mkdir("AutoNetT")

    def is_Admin(self, user):
        try:
            if Configuration.enable_session:
                conn_logs = sqlite3.connect(self.database_path)
                curs_logs = conn_logs.cursor()
                curs_logs.execute("SELECT USER_ID FROM ADMINS WHERE USER_ID=? AND USER_ID", (user, user))
                results = curs_logs.fetchone()
                if results is not None:
                    return True
                return False
            return True
        except Exception as err:
            print("Administrator rights required to run this software: " + str(err))

    def coordinator(self, host):
        work_sheet = topology_Map
        Telnet.open(host, port=23)
        self.validate_IPv4(host())
        if is_Admin(self.user):
            return True
        else:
            return False

    @staticmethod
    def validate_IPv4(ipv4addr):
        split_ipv4addr = ipv4addr.split(".")
        conv_ipaddr = []
        if len(split_ipv4addr) > 4:
            return f'Invalid IPv4 address'
        else:
            for i in split_ipv4addr:
                i = int(i)
                if i < 0 or i > 255:
                    print("Invalid IPv4 address")
                else:
                    conv_ipaddr.append(i)
        return ipv4addr

    def topology_Map(self):
        work_sheet = Workbook()
        work_sheet.title = "Test Output"
        work_sheet.cell(row=1, column=1, value="Device Type")
        work_sheet.cell(row=1, column=2, value="Device Name")
        work_sheet.cell(row=1, column=3, value="Interface")
        work_sheet.cell(row=1, column=4, value="IP Address")
        work_sheet.cell(row=1, column=5, value="Active Protocols")
        self.workbook.save("AutoNetT_Top_Map_Output.xlsx")
        return work_sheet

    def network_Search(self, file, workbook):
        file.open("AutoNetT_Top_Map_Output.xlsx", "w")
        sheet = workbook.active
        sheet['D1'] = self.host
        if is_Admin():
            try:
                router = ConnectHandler(router_type='cisco_ios', ip=host, username=input(), password=input())
                output = router.send_command("enable", "configure terminal", "cdp enable", "show run")
                print(output)
            except Exception as err:
                print("Connection to router failed: " + str(err))
            pass
